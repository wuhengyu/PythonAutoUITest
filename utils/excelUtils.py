# -*- coding: utf-8 -*-
# Time    : 2023/3/19 14:46
# Author  : Walter
# File    : excelUtils.py
# License : (C)Copyright Walter
# Version : 1.0
# Desc    :

import openpyxl

# class ExcelHandler:
#     def __init__(self, file_path):
#         self.file_path = file_path
#         self.workbook = openpyxl.load_workbook(file_path)
#         self.sheet = self.workbook.active
#
#     def get_cell_value(self, row, column):
#         return self.sheet.cell(row=row, column=column).value
#
#     def set_cell_value(self, row, column, value):
#         self.sheet.cell(row=row, column=column).value = value
#
#     def save_changes(self):
#         self.workbook.save(self.file_path)


class ExcelTool:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)

    def get_sheet(self, sheet_name):
        return self.workbook[sheet_name]

    def get_cell_value(self, sheet_name, row, column):
        sheet = self.get_sheet(sheet_name)
        return sheet.cell(row=row, column=column).value

    def set_cell_value(self, sheet_name, row, column, value):
        sheet = self.get_sheet(sheet_name)
        sheet.cell(row=row, column=column).value = value

    def save(self):
        self.workbook.save(self.file_path)

excelHandler = ExcelTool("../files/testExcel.xlsx")
value = excelHandler.get_cell_value("Sheet1", 1, 1)
print(value)
